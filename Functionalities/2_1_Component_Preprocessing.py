
# --------------------------------------------------------------------------------------------------
# ==================================================================================================
# Title:        GDP Component Pre-Processing
#
# Author:       ChatGPT 5.2, prompted and ammended by Jan Ole Westphal
#
# Description:  Takes the ifo forecast overview excel and extracts vintage tables for all compontents
#
# ==================================================================================================
# --------------------------------------------------------------------------------------------------




import os, sys
import openpyxl, re, datetime as dt
from openpyxl.utils import get_column_letter




# ==================================================================================================
#                                      DEFINE FOLDER STRUCTURE
# ==================================================================================================

# Ensure project root is in sys.path

wd = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if wd not in sys.path:
    sys.path.insert(0, wd)

# Dynamically find the latest 'ifo_Konjunkturprognose' Excel file
ifo_qoq_input_path = os.path.join(wd, '0_0_Data', '0_Forecast_Inputs')
excel_files = [f for f in os.listdir(ifo_qoq_input_path) if f.endswith('.xlsx') and f.startswith('ifo_Konjunkturprognose')]

if not excel_files:
    raise ValueError(f"No Excel files starting with 'ifo_Konjunkturprognose' found in the directory {ifo_qoq_input_path}.")

def _season_year_key(filename):
    last_part = filename.rsplit('_', 1)[-1]
    last_part = last_part.replace('.xlsx', '')
    if len(last_part) < 3:
        return (0, 0)
    season = last_part[0]
    year = int(last_part[1:])
    season_order = {'W': 1, 'H': 2, 'S': 3, 'F': 4}
    return (year, season_order.get(season, 0))

excel_files_sorted = sorted(excel_files, key=_season_year_key, reverse=True)
SRC_PATH = os.path.join(ifo_qoq_input_path, excel_files_sorted[0])
wb_src = openpyxl.load_workbook(SRC_PATH, data_only=True)

OUT_PATH = os.path.join(wd, '0_0_Data', '0_Forecast_Inputs','1_ifo_quarterly_components', "ifo_BIP_Komponenten.xlsx")





print("\n\nCreating component-level vintage dataset from ifo quarterly excel...\n")


# -------------------------------------------------------------------------------------------------#
# =================================================================================================#
#                                       PROCESSING PIPELINE                                        #
# =================================================================================================#
# -------------------------------------------------------------------------------------------------#


# ==================================================================================================
#                                          FUNCTIONALITIES
# ==================================================================================================

ROMAN_TO_Q = {"I": 1, "II": 2, "III": 3, "IV": 4}
PUB_LETTER_TO_Q = {"F": 1, "S": 2, "H": 3, "W": 4}

# Output sheet mapping: abbreviation -> list of row-label patterns (normalised substrings)
VAR_MAP = {
    "GDP":     ["bruttoinlandsprodukt"],
    "PRIVCON": ["private konsumausgaben", "privater konsum", "privater verbrauch"],
    "PUBCON":  ["konsumausgaben des staates", "offentlicher konsum", "konsum des staates"],
    "CONSTR":  ["bauten"],
    "OPA":     ["sonstige anlagen"],
    "EQUIPMENT": ["ausrustungen", "ausrüstungen", "ausruestungen", "ausrustungsinvestitionen", 
                  "ausruestungsinvestitionen", "ausrüstungsinvestitionen" ],
    "INVINV":  ["vorratsinvestitionen", "vorratsveraenderungen", "vorratsveranderungen", "vorrate"],
    "DOMUSE":  ["inlandische verwendung", "inlandsnachfrage", "inlaendische verwendung"],
    "TRDBAL":  ["aussenbeitrag", "außenbeitrag"],
    "EXPORT":  ["exporte"],
    "IMPORT":  ["importe"],
}

def norm(s: str) -> str:
    s = (s or "").strip().lower()
    trans = str.maketrans({"ä": "a", "ö": "o", "ü": "u", "ß": "ss"})
    s = s.translate(trans)
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def quarter_midmonth_date(year: int, q: int) -> dt.datetime:
    # same convention as the existing BIP sheet: Q1->Feb 15, Q2->May 15, Q3->Aug 15, Q4->Nov 15
    month = {1: 2, 2: 5, 3: 8, 4: 11}[q]
    return dt.datetime(year, month, 15)

def match_variable(label_norm: str):
    for var, patterns in VAR_MAP.items():
        for p in patterns:
            if p in label_norm:
                return var
    return None

def read_publication_dates_from_BIP(wb):
    """
    Map (letter, year) -> datenstand datetime from sheet BIP:
      row1: letter (W/F/S/H)
      row2: year (YYYY)
      row3: datenstand (datetime)
    """
    ws = wb["BIP"]
    pub_dates = {}
    for c in range(2, ws.max_column + 1):
        letter = ws.cell(1, c).value
        year = ws.cell(2, c).value
        dts = ws.cell(3, c).value
        if isinstance(letter, str) and isinstance(year, (int, float)) and isinstance(dts, dt.datetime):
            pub_dates[(letter.strip(), int(year))] = dts
    return pub_dates

def find_first_table_header(ws, search_rows=80, search_cols=60):
    """
    Find (year_row, roman_row) for the FIRST table only.
    - year_row has >=2 integer years
    - roman_row is the next row below containing multiple roman numerals
    """
    for r in range(1, search_rows + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, search_cols + 1)]
        years = [v for v in row_vals if isinstance(v, (int, float)) and int(v) == v and 1900 <= int(v) <= 2100]
        if len(years) >= 2:
            for rr in range(r + 1, min(r + 10, search_rows) + 1):
                rv = [ws.cell(rr, c).value for c in range(1, search_cols + 1)]
                romans = [v for v in rv if isinstance(v, str) and v.strip() in ROMAN_TO_Q]
                if len(romans) >= 2:
                    return r, rr
    raise ValueError(f"Could not find table header in sheet '{ws.title}'")

def parse_sheet_table(ws):
    """
    Parse one quarterly sheet (e.g. H25) and return:
      pub_key = (letter, year)
      series_by_label = {normalised_row_label: {(target_year, target_q): value}}
    Only the first table is used (if stacked tables exist).
    Empty rows (pre-2016) are ignored.
    """
    m = re.fullmatch(r"([FSHW])(\d{2})", ws.title.strip())
    if not m:
        raise ValueError(f"Unexpected sheet name: {ws.title}")
    pub_letter, yy = m.group(1), int(m.group(2))
    pub_year = 2000 + yy if yy <= 79 else 1900 + yy
    pub_key = (pub_letter, pub_year)

    year_row, roman_row = find_first_table_header(ws)

    # propagate years across columns
    years = []
    last_year = None
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(year_row, c).value
        if isinstance(v, (int, float)) and int(v) == v and 1900 <= int(v) <= 2100:
            last_year = int(v)
        years.append(last_year)

    # target periods per column
    target_periods = {}  # col -> (year,q)
    for c in range(1, max_col + 1):
        v = ws.cell(roman_row, c).value
        if isinstance(v, str):
            vv = v.strip()
            if vv in ROMAN_TO_Q and years[c - 1] is not None:
                target_periods[c] = (years[c - 1], ROMAN_TO_Q[vv])

    series_by_label = {}
    started = False

    for r in range(roman_row + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value

        # stop if a second table header appears later (stacked tables)
        row_vals = [ws.cell(r, c).value for c in range(1, min(max_col, 60) + 1)]
        year_ints = [v for v in row_vals if isinstance(v, (int, float)) and int(v) == v and 1900 <= int(v) <= 2100]
        if started and len(year_ints) >= 2:
            break

        if isinstance(a, str):
            label_raw = a.strip()

            # stop at footnotes/source lines
            if norm(label_raw).startswith(("a ", "quelle", "source", "anmerk")):
                break

            vals = {}
            any_num = False
            for c, (yy, qq) in target_periods.items():
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    any_num = True
                    vals[(yy, qq)] = float(v)

            if any_num:
                started = True
                k = norm(label_raw)
                series_by_label.setdefault(k, {}).update(vals)

        # ignore empty rows; they occur pre-2016

    return pub_key, series_by_label

def sort_pub_key(pub_key):
    letter, year = pub_key
    return (year, PUB_LETTER_TO_Q[letter])

def sheetname_sort_key(name):
    m = re.fullmatch(r"([FSHW])(\d{2})", name)
    letter, yy = m.group(1), int(m.group(2))
    year = 2000 + yy if yy <= 79 else 1900 + yy
    return (year, PUB_LETTER_TO_Q[letter])










# ==================================================================================================
#                                       EXECUTION PIPELINE
# ==================================================================================================


def main():
    wb_src = openpyxl.load_workbook(SRC_PATH, data_only=True)
    pub_dates = read_publication_dates_from_BIP(wb_src)

    # quarterly sheets are dynamic: any sheet matching [FSHW]\d{2} is included
    quarterly_sheets = sorted(
        [s for s in wb_src.sheetnames if re.fullmatch(r"[FSHW]\d{2}", s)],
        key=sheetname_sort_key,
    )

    # data[var][pub_key][target]=(value)
    data = {var: {} for var in VAR_MAP}
    all_targets = set()
    all_pubs = set()

    for sh in quarterly_sheets:
        ws = wb_src[sh]
        pub_key, series_by_label = parse_sheet_table(ws)
        all_pubs.add(pub_key)

        for lbl, series in series_by_label.items():
            var = match_variable(lbl)
            if var:
                data[var].setdefault(pub_key, {}).update(series)
                all_targets |= set(series.keys())

    # sorted target quarters
    all_targets_sorted = sorted(all_targets, key=lambda t: (t[0], t[1]))
    target_dates = [quarter_midmonth_date(y, q) for (y, q) in all_targets_sorted]

    # sorted publication columns
    all_pubs_sorted = sorted(all_pubs, key=sort_pub_key)
    pub_letters, pub_years, pub_datetimes = [], [], []
    for (letter, year) in all_pubs_sorted:
        pub_letters.append(letter)
        pub_years.append(year)
        pub_datetimes.append(pub_dates.get((letter, year), quarter_midmonth_date(year, PUB_LETTER_TO_Q[letter])))

    # create output workbook
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for var in VAR_MAP.keys():
        ws_out = wb_out.create_sheet(var)

        # Header like BIP
        ws_out.cell(1, 1).value = "Prognose"
        ws_out.cell(2, 1).value = "Jahr"
        ws_out.cell(3, 1).value = "Datenstand"

        for j, (letter, year, dts) in enumerate(zip(pub_letters, pub_years, pub_datetimes), start=2):
            ws_out.cell(1, j).value = letter
            ws_out.cell(2, j).value = year
            ws_out.cell(3, j).value = dts

        # Target dates as index
        for i, td in enumerate(target_dates, start=4):
            ws_out.cell(i, 1).value = td

        # Fill values
        target_index = {t: idx for idx, t in enumerate(all_targets_sorted)}  # (year,q)->0-based
        for j, pub_key in enumerate(all_pubs_sorted, start=2):
            series = data[var].get(pub_key, {})
            for t, val in series.items():
                r = 4 + target_index[t]
                ws_out.cell(r, j).value = val

        # Formatting (optional)
        ws_out.freeze_panes = "B4"
        date_fmt = "yyyy-mm-dd"
        ws_out.column_dimensions["A"].width = 14
        for r in range(3, 4 + len(target_dates)):
            ws_out.cell(r, 1).number_format = date_fmt
        for c in range(2, 2 + len(pub_datetimes)):
            ws_out.cell(3, c).number_format = date_fmt
            if c < 60:
                ws_out.column_dimensions[get_column_letter(c)].width = 11


    # ------------------------------------------------------------------
    # Delete Faulty Cells
    # ------------------------------------------------------------------
    delete_rows = {
        dt.datetime(2015, 2, 15),
        dt.datetime(2015, 5, 15),
        dt.datetime(2015, 8, 15),
        dt.datetime(2015, 11, 15),
    }

    delete_cols = {
        dt.datetime(2017, 11, 14),
        dt.datetime(2018, 2, 14),
        dt.datetime(2018, 5, 15),
        dt.datetime(2018, 8, 14),
        dt.datetime(2018, 11, 14),
    }

    for ws in wb_out.worksheets:
        # map row dates
        row_map = {
            ws.cell(r, 1).value: r
            for r in range(4, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, dt.datetime)
        }

        # map column dates
        col_map = {
            ws.cell(3, c).value: c
            for c in range(2, ws.max_column + 1)
            if isinstance(ws.cell(3, c).value, dt.datetime)
        }

        for rd in delete_rows:
            if rd in row_map:
                r = row_map[rd]
                for cd in delete_cols:
                    if cd in col_map:
                        ws.cell(r, col_map[cd]).value = None

    wb_out.save(OUT_PATH)

if __name__ == "__main__":
    main()




# --------------------------------------------------------------------------------------------------
print(f"\nComponent-Level Data Preprocessing complete \n")
# --------------------------------------------------------------------------------------------------




# -------------------------------------------------------------------------------------------------#
# =================================================================================================#
#                                        End of Code                                               #
# =================================================================================================#
# -------------------------------------------------------------------------------------------------#